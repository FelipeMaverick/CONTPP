from flask import Flask, request, send_file, jsonify, send_from_directory
from flask_cors import CORS
import os, io, zipfile, re, tempfile
import openpyxl

app = Flask(__name__, static_folder='.')
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')
os.makedirs(TEMPLATES_DIR, exist_ok=True)

FIELD_ORDER = {
    'contrato_talent': [
        "contratada_nome", "contratada_cpf", "contratada_rg", "contratada_endereco",
        "agencia_nome", "agencia_endereco", "agencia_cnpj",
        "cliente", "marca", "titulo_campanha", "formato_campanha",
        "exclusividade", "funcao_ator_modelo", "diarias_producao",
        "valor_bruto_total_brl", "valor_contratada_brl", "taxa_agenciamento_brl",
        "valor_diaria_extra_brl", "valor_hora_extra_brl",
        "valor_bruto_total_eur", "valor_contratada_eur", "taxa_agenciamento_eur",
        "valor_diaria_extra_eur", "valor_hora_extra_eur", "prazo_forma_pagamento",
        "contratada_nome", "contratada_nacionalidade", "contratada_cpf",
        "contratada_rg", "contratada_endereco",
        "agencia_nome", "agencia_endereco", "agencia_cnpj",
        "cliente", "marca", "titulo_campanha", "formato_campanha",
        "exclusividade", "funcao_ator_modelo", "diarias_producao",
        "valor_bruto_total_brl", "valor_contratada_brl", "taxa_agenciamento_brl",
        "valor_diaria_extra_brl", "valor_hora_extra_brl",
        "valor_bruto_total_eur", "valor_contratada_eur", "taxa_agenciamento_eur",
        "valor_diaria_extra_eur", "valor_hora_extra_eur", "prazo_forma_pagamento",
    ]
}

def find_template(tipo):
    """Find template file, checking templates/ folder and root."""
    paths = [
        os.path.join(TEMPLATES_DIR, f'{tipo}.docx'),
        os.path.join(BASE_DIR, f'{tipo}.docx'),
        os.path.join(TEMPLATES_DIR, 'contrato_talent.docx'),
        os.path.join(BASE_DIR, 'contrato_talent.docx'),
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return None

def fill_contract(template_path, data, field_order):
    with open(template_path, 'rb') as f:
        template_bytes = f.read()

    with tempfile.TemporaryDirectory() as tmpdir:
        docx_path = os.path.join(tmpdir, 'contract.docx')
        with open(docx_path, 'wb') as f:
            f.write(template_bytes)

        extract_dir = os.path.join(tmpdir, 'extracted')
        with zipfile.ZipFile(docx_path, 'r') as z:
            z.extractall(extract_dir)

        doc_xml_path = os.path.join(extract_dir, 'word', 'document.xml')
        with open(doc_xml_path, 'r', encoding='utf-8') as f:
            xml = f.read()

        field_idx = 0
        def replace_next(m):
            nonlocal field_idx
            if field_idx < len(field_order):
                key = field_order[field_idx]
                value = str(data.get(key, '') or '')
                field_idx += 1
                return f'<w:t>{value}</w:t>'
            return m.group(0)

        xml = re.sub(r'<w:t>\[\.\]</w:t>', replace_next, xml)

        with open(doc_xml_path, 'w', encoding='utf-8') as f:
            f.write(xml)

        output_path = os.path.join(tmpdir, 'output.docx')
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, extract_dir)
                    zout.write(file_path, arcname)

        with open(output_path, 'rb') as f:
            return f.read()


@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


@app.route('/templates', methods=['GET'])
def list_templates():
    templates = []
    seen = set()
    for folder in [TEMPLATES_DIR, BASE_DIR]:
        for f in os.listdir(folder):
            if f.endswith('.docx') and f not in seen:
                seen.add(f)
                key = f.replace('.docx', '')
                label = key.replace('_', ' ').title()
                templates.append({'key': key, 'label': label, 'file': f})
    return jsonify(templates)


@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    tipo = data.get('tipo', 'contrato_talent')
    template_file = find_template(tipo)

    if not template_file:
        return jsonify({'error': 'Template não encontrado'}), 404

    field_order = FIELD_ORDER.get(tipo, FIELD_ORDER['contrato_talent'])

    try:
        docx_bytes = fill_contract(template_file, data, field_order)
        name = (data.get('contratada_nome') or 'contrato').replace(' ', '_')
        return send_file(
            io.BytesIO(docx_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            as_attachment=True,
            download_name=f'Contrato_{name}.docx'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate-batch', methods=['POST'])
def generate_batch():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    tipo = request.form.get('tipo', 'contrato_talent')
    template_file = find_template(tipo)
    if not template_file:
        return jsonify({'error': 'Template não encontrado'}), 404

    field_order = FIELD_ORDER.get(tipo, FIELD_ORDER['contrato_talent'])
    wb = openpyxl.load_workbook(request.files['file'])
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]

    if not rows:
        return jsonify({'error': 'Nenhuma linha encontrada'}), 400

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, row in enumerate(rows):
            row_data = {k: str(v) if v is not None else '' for k, v in row.items()}
            docx_bytes = fill_contract(template_file, row_data, field_order)
            name = row_data.get('contratada_nome', f'talento_{i+1}').replace(' ', '_')
            zf.writestr(f'Contrato_{name}.docx', docx_bytes)

    zip_buffer.seek(0)
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name='Contratos.zip')


@app.route('/upload-template', methods=['POST'])
def upload_template():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo'}), 400
    f = request.files['file']
    filename = f.filename.replace(' ', '_').lower()
    if not filename.endswith('.docx'):
        return jsonify({'error': 'Apenas .docx'}), 400
    f.save(os.path.join(TEMPLATES_DIR, filename))
    key = filename.replace('.docx', '')
    FIELD_ORDER[key] = FIELD_ORDER['contrato_talent']
    return jsonify({'ok': True, 'key': key, 'label': key.replace('_', ' ').title()})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port)
